# Script per fer el càlcul de les hores de Gladiator i de producció AIRA per cada mes
# He de generar un excel que tingui 2 fulles
# A la primera fulla he de copiar La part de baix del control de marcatges per cada dia!
# Es tracta de fer tot el llistat de control de marcatges del mes, i anar dia a dia i copiar la part de baix del GP5 a l'excel
# A la segona fulla hi ha d'haver el resum de tot el mes per comprovar la suma tota de les hores
# En aquesta segona fulla he de crear les columnes R, S i T
# En la R he de posar C o A depenent de si el marcatge és de Casa o de AIRA
# Les columnes S i T les puc copiar de la fulla de mostra C:\Scripts\hores_marc.xlsx


import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Carregar l'arxiu Excel
df = pd.read_excel("C:\\Scripts\\hores_marc.xlsx")

# Eliminar la primera columna
df = df.iloc[:, 1:]

# Eliminar files amb valors concrets
df = df[~df.iloc[:, 0].isin(["Ordre Fabricació", "22/3396"])]

# Eliminar columnes específiques
cols_to_drop = df.columns[[2, 3, 8, 9, 10, 11, 12, 13, 14, 15]]
df.drop(columns=cols_to_drop, inplace=True)

# Calcular sumes i percentatges
col_ordre = df.columns[0]
col_total_hores = df.columns[3]

suma_0016 = df[df[col_ordre] == "25/0016"][col_total_hores].sum()
suma_1604 = df[df[col_ordre] == "25/1604"][col_total_hores].sum()
suma_total = df[col_total_hores].sum()

percent_0016 = suma_0016 / suma_total if suma_total != 0 else 0
percent_1604 = suma_1604 / suma_total if suma_total != 0 else 0

# Guardar resultat netejat
output_path = r"C:\Scripts\resultat_netejat.xlsx"
df.to_excel(output_path, index=False, sheet_name='Resultat')

# Tornar a obrir per aplicar formats
wb = load_workbook(output_path)
ws = wb['Resultat']

# Escriure dades
ws['I2'] = suma_0016
ws['I3'] = suma_1604
ws['H2'] = f"{percent_0016:.2%}"
ws['H3'] = f"{percent_1604:.2%}"
ws['J2'] = "Total Hores Producció AIRA"
ws['J3'] = "Total Hores Projecte GLADIATOR"

# Ajustar amplades automàticament
for col_idx, col in enumerate(ws.columns, 1):
    max_length = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
    ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2

# Aplicar estil a columnes H, I, J (8, 9, 10)
font_vermella_negreta = Font(color="FF0000", bold=True)

for col_idx in [8, 9, 10]:
    col_letter = get_column_letter(col_idx)
    for row in range(1, ws.max_row + 1):
        cell = ws[f"{col_letter}{row}"]
        cell.font = font_vermella_negreta
        
# Guardar els canvis
wb.save(output_path)
